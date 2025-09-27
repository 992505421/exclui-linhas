from typing import List, Tuple
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
import io
import re

def row_matches(ws: Worksheet, row_idx: int, patterns: List[str]) -> bool:
    r = re.compile("|".join(map(re.escape, patterns)), re.IGNORECASE)
    for cell in ws[row_idx]:
        if cell.value is not None and r.search(str(cell.value)):
            return True
    return False

def process_excel(content: bytes, original_name: str, names: List[str]) -> Tuple[bytes, str]:
    wb = load_workbook(io.BytesIO(content))
    ws = wb.active  # pega a primeira aba

    to_delete = []
    for i in range(ws.max_row, 0, -1):  # varre de baixo pra cima
        if row_matches(ws, i, names):
            to_delete.append(i)
    for i in to_delete:
        ws.delete_rows(i, 1)

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    new_name = f"{original_name.rsplit('.',1)[0]}_limpo.xlsx"
    return out.read(), new_name
